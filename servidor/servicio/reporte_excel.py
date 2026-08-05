import os
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "recursos", "logo.png")

_AZUL = "1E3A8A"
_GRIS_CLARO = "F1F5F9"


def _preparar_hoja(wb: Workbook, titulo_hoja: str, titulo: str, subtitulo: str, num_columnas: int):
    ws = wb.active
    ws.title = titulo_hoja

    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width, img.height = 150, 45
        ws.add_image(img, "A1")

    fila_titulo = 4
    ws.cell(row=fila_titulo, column=1, value=titulo).font = Font(size=14, bold=True, color=_AZUL)
    ws.cell(row=fila_titulo + 1, column=1, value=subtitulo).font = Font(size=10, italic=True, color="475569")
    return ws, fila_titulo + 3  # fila donde empieza la tabla de encabezados


def _escribir_tabla(ws, fila_inicio: int, encabezados: list, filas: list):
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_inicio, column=col, value=texto)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center")

    for i, fila in enumerate(filas):
        r = fila_inicio + 1 + i
        for col, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=col, value=valor)
            if i % 2 == 1:
                celda.fill = PatternFill("solid", fgColor=_GRIS_CLARO)

    for col in range(1, len(encabezados) + 1):
        letra = get_column_letter(col)
        max_len = max([len(str(encabezados[col - 1]))] + [len(str(f[col - 1])) for f in filas] or [10])
        ws.column_dimensions[letra].width = min(max(max_len + 2, 12), 40)

    return fila_inicio + 1 + len(filas)


def generar_excel_ventas(datos: dict) -> BytesIO:
    wb = Workbook()
    subtitulo = f"Ventas del {datos['desde']} al {datos['hasta']} · generado {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws, fila = _preparar_hoja(wb, "Ventas", "Reporte de ventas", subtitulo, 7)

    encabezados = ["Fecha", "Venta", "Vendedor", "Producto", "Cantidad", "Precio unitario", "Subtotal"]
    filas = [[
        f["fecha"], f["venta_id"], f["vendedor"], f["producto"],
        f["cantidad"], f["precio_unitario"], f["subtotal"],
    ] for f in datos["filas"]]

    fila_fin = _escribir_tabla(ws, fila, encabezados, filas)

    ws.cell(row=fila_fin + 2, column=1, value="Total de ventas:").font = Font(bold=True)
    ws.cell(row=fila_fin + 2, column=2, value=datos["num_ventas"])
    ws.cell(row=fila_fin + 3, column=1, value="Monto total:").font = Font(bold=True)
    ws.cell(row=fila_fin + 3, column=2, value=datos["total_general"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_mas_vendidos(datos: dict) -> BytesIO:
    wb = Workbook()
    subtitulo = f"Periodo {datos['desde']} al {datos['hasta']} · generado {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws, fila = _preparar_hoja(wb, "Mas vendidos", "Productos más vendidos", subtitulo, 4)

    encabezados = ["#", "Producto", "Unidades vendidas", "Total vendido"]
    filas = [[i, f["producto"], f["unidades_vendidas"], f["total_vendido"]]
             for i, f in enumerate(datos["filas"], start=1)]

    _escribir_tabla(ws, fila, encabezados, filas)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
